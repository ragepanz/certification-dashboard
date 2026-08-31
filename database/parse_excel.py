import openpyxl
import datetime
import re
import json
import os
import warnings
warnings.filterwarnings('ignore')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'Training Dinas TN.xlsx')
OUTPUT_PATH = os.path.join(BASE_DIR, 'database', 'excel_import.json')
MATRIX_PATH = os.path.join(BASE_DIR, 'database', 'training_man_matrix.json')

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_data = wb['TrainingData']
ws_man = wb['TrainingMan'] if 'TrainingMan' in wb.sheetnames else None

# Build or load TrainingMan matrix
matrix = {}
if ws_man:
    for r in range(2, ws_man.max_row + 1):
        job = ws_man.cell(r, 2).value
        code = ws_man.cell(r, 3).value
        tr = ws_man.cell(r, 4).value
        val = ws_man.cell(r, 6).value
        nn = ws_man.cell(r, 8).value
        if job and tr:
            j_key = str(job).strip().lower()
            t_key = str(tr).strip().lower()
            if j_key not in matrix:
                matrix[j_key] = {}
            matrix[j_key][t_key] = {
                'validity': str(val).strip() if val else 'Forever',
                'no_need': True if str(nn).strip().lower() == 'yes' else False,
                'code': str(code).strip() if code else None
            }
    with open(MATRIX_PATH, 'w', encoding='utf-8') as f:
        json.dump(matrix, f, indent=2)

headers = {c: ws_data.cell(2, c).value for c in range(1, ws_data.max_column + 1)}

EXPANDING_COLS = 68

def parse_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        try:
            dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(v))
            return dt.strftime('%Y-%m-%d')
        except:
            pass
    s = str(v).strip()
    m1 = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m1:
        try:
            return f'{int(m1.group(1)):04d}-{int(m1.group(2)):02d}-{int(m1.group(3)):02d}'
        except:
            pass
    return None

def is_valid_training(v):
    s = str(v).strip().upper() if v else ''
    return s in ('VALID', 'EXPIRING', 'EXPIRED')

data = []
for row_idx in range(3, ws_data.max_row + 1):
    emp_id = ws_data.cell(row_idx, 2).value
    emp_name = ws_data.cell(row_idx, 3).value
    job_title = ws_data.cell(row_idx, 4).value
    unit = ws_data.cell(row_idx, 6).value or ws_data.cell(row_idx, 7).value or 'TN'
    
    if not (emp_id and emp_name):
        continue
    
    emp_id_str = str(emp_id).strip()
    emp_name_str = str(emp_name).strip()
    job_title_str = str(job_title).strip() if job_title else None
    unit_str = str(unit).strip()
    
    j_key = job_title_str.lower() if job_title_str else 'unknown'
    
    certs = []
    
    # Kolom 08-26: paired date + status columns (step 2)
    for c in range(8, 27, 2):
        if c > EXPANDING_COLS:
            break
        cert_name = headers.get(c)
        status_col = c + 1
        if not cert_name:
            continue
        
        date_val = ws_data.cell(row_idx, c).value
        status_val = ws_data.cell(row_idx, status_col).value
        
        parsed_date = parse_date(date_val)
        status_str = str(status_val).strip().upper() if status_val else ''
        
        if not is_valid_training(status_val):
            continue
        
        if parsed_date:
            t_key = cert_name.strip().lower()
            
            # Cek terhadap matriks TrainingMan
            is_two_year = True
            is_no_need = False
            
            if j_key in matrix and t_key in matrix[j_key]:
                rule = matrix[j_key][t_key]
                if rule['no_need']:
                    is_no_need = True
                elif rule['validity'] == 'Forever':
                    is_two_year = False
                else:
                    is_two_year = True
            elif cert_name == 'Dangerous Good':
                is_two_year = False
            
            if is_no_need or not is_two_year:
                # Modul ini untuk Job Title bersangkutan adalah Permanen (Forever)
                certs.append({
                    'name': cert_name,
                    'expiry_date': None,
                    'issue_date': parsed_date,
                    'status': 'valid',
                    'is_periodic': False
                })
            else:
                # Modul 2-Year berkala
                certs.append({
                    'name': cert_name,
                    'expiry_date': parsed_date,
                    'issue_date': None,
                    'status': status_str.lower(),
                    'is_periodic': True
                })
    
    # Kolom 28-68: single date only = Permanen (Forever)
    for c in range(28, EXPANDING_COLS + 1):
        cert_name = headers.get(c)
        date_val = ws_data.cell(row_idx, c).value
        
        if not cert_name:
            continue
        
        if any(cert['name'] == cert_name for cert in certs):
            continue
        
        parsed_date = parse_date(date_val)
        
        if parsed_date:
            certs.append({
                'name': cert_name,
                'expiry_date': None,
                'issue_date': parsed_date,
                'status': 'valid',
                'is_periodic': False
            })
    
    if certs:
        data.append({
            'employee_number': emp_id_str,
            'name': emp_name_str,
            'job_title': job_title_str,
            'unit': unit_str,
            'certs': certs,
        })

total = sum(len(d['certs']) for d in data)
periodic_total = sum(sum(1 for c in d['certs'] if c.get('is_periodic')) for d in data)
forever_total = total - periodic_total

print(f"Parsed {len(data)} employees with {total} certifications.")
print(f"Periodic (2-Year) certs: {periodic_total}, Forever (Permanent) certs: {forever_total}")

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("Saved to database/excel_import.json")